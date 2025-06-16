from PyQt5.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QComboBox,
    QDateEdit,
    QPushButton,
    QListWidget,
    QFileDialog,
    QMessageBox,
    QStackedWidget,
    QSizePolicy,
    QListWidgetItem,
)
from PyQt5.QtCore import QDate
from PyQt5.QtGui import QFont
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.dates as mdates
import matplotlib.ticker as ticker
from datetime import datetime, timedelta
from matplotlib.backends.backend_pdf import PdfPages
from _sopian.path_utils import get_database_path
import pandas as pd
from openpyxl import Workbook 
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import json


import json  # ADD: Import json module

class HistoryManager:
    @staticmethod
    def load_history(username, start_time, deadline, status_filter="all"):
        done = {}
        failed = {}
        entries = []
        
        # CHANGE: Use JSON file instead of text file
        history_file = get_database_path("history.json")
        try:
            with open(history_file, "r") as f:
                data = json.load(f)
                for entry in data["history"]:  # Process JSON array
                    # Map JSON fields to variables (key change: 'name' -> 'task')
                    task = entry.get("name", "")
                    description = entry.get("description", "")
                    start_time = entry.get("start_time", "")
                    deadline_entry = entry.get("deadline", "")
                    priority = entry.get("priority", "")
                    reminder = entry.get("reminder", "")
                    status = entry.get("status", "")
                    schedule = entry.get("schedule", "")
                    entry_username = entry.get("username", "")

                    # Rest of the logic remains the same from original code
                    if entry_username != username:
                        continue

                    # Original date parsing logic
                    if "on" in status:
                        date_part = status.split("on ")[-1].strip()
                        try:
                            status_date = datetime.strptime(date_part, "%Y-%m-%d").date()
                        except ValueError:
                            continue
                    else:
                        try:
                            status_date = datetime.strptime(status, "%Y-%m-%d").date()
                        except ValueError:
                            continue
                    
                    # Rest of original processing logic
                    status_type = "failed" if "failed" in status.lower() else "done"
                    if status_filter not in ["all", status_type]:
                        continue

                    date_str = status_date.strftime("%Y-%m-%d")
                    if status_type == "failed":
                        failed[date_str] = failed.get(date_str, 0) + 1
                    else:
                        done[date_str] = done.get(date_str, 0) + 1
                    
                    entries.append({
                        'task': task,  # Mapped from JSON 'name' field
                        'description': description,
                        'start_time': start_time,
                        'deadline': deadline_entry,
                        'priority': priority,
                        'reminder': reminder,
                        'status': status,
                        'schedule': schedule,
                        'username': entry_username
                    })
        except (FileNotFoundError, json.JSONDecodeError, KeyError):
            pass
        return done, failed, entries


class HistoryWidget(QWidget):
    def __init__(self, username):
        super().__init__()
        self.username = username  # NEW: Store username
        self.initUI()
        self.load_initial_data()

    def load_initial_data(self):
        """Load data with default date range on first open"""
        self.update_date_range()  # Sets default dates
        self.update_display()

    def initUI(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 10, 20, 10)
        main_layout.setSpacing(8)

        # Header remains the same
        header = QHBoxLayout()
        title = QLabel("History")
        title.setFont(QFont("Arial", 20, QFont.Bold))
        title.setStyleSheet("margin: 0;")
        header.addWidget(title)
        header.addStretch()
        main_layout.addLayout(header)

        # Controls container
        controls_container = QWidget()
        controls_layout = QHBoxLayout(controls_container)
        controls_layout.setContentsMargins(0, 0, 0, 0)
        controls_layout.setSpacing(10)

        # Initialize control widgets
        self.range_combo = QComboBox()

        self.status_combo = QComboBox()

        self.view_combo = QComboBox()

        self.start_date = QDateEdit(calendarPopup=True)

        self.end_date = QDateEdit(calendarPopup=True)
        
        self.export_btn = QPushButton("Export")

        # Update 19 Mei 2025
        # Export format combo
        self.export_format_combo = QComboBox()
        self.export_format_combo.addItems(["PDF", "Excel"])

        # Add controls to layout
        controls_layout.addWidget(self.range_combo)
        controls_layout.addWidget(self.status_combo)

        controls_layout.addWidget(QLabel("View:"))
        controls_layout.addWidget(self.view_combo)

        controls_layout.addWidget(QLabel("From:"))
        controls_layout.addWidget(self.start_date)

        controls_layout.addWidget(QLabel("To:"))
        controls_layout.addWidget(self.end_date)

        controls_layout.addWidget(self.export_format_combo)
        controls_layout.addWidget(self.export_btn)
        

        main_layout.addWidget(controls_container)

        # View stack
        self.stacked_widget = QStackedWidget()
        self.stacked_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # Initialize views
        self.figure = Figure(figsize=(10, 5), tight_layout=True)
        self.canvas = FigureCanvas(self.figure)

        # Configure list style
        self.history_list = QListWidget()
        self.history_list.setStyleSheet(
            """
            QListWidget {
                background: white;
                border-radius: 8px;
                padding: 8px;
            }
        """
        )   

        # Add views to stack
        self.stacked_widget.addWidget(self.canvas)
        self.stacked_widget.addWidget(self.history_list)

        main_layout.addWidget(self.stacked_widget, 1)

        # Configure control widgets
        self._configure_controls()

        # Initial setup
        self.update_date_range()
        self.toggle_view()

    def _configure_controls(self):
        # Combo box items
        self.range_combo.addItems(["This Week", "This Month", "This Year", "Custom"])

        self.status_combo.addItems(["All", "Done", "Failed"])

        # UPDATED: Add new view options
        self.view_combo.addItems(["Bar Graph", "Line Graph", "Pie Chart", "Text View"])

        # Style controls
        control_style = "padding: 5px; border-radius: 5px;"
        for widget in [
            self.range_combo,
            self.status_combo,
            self.view_combo,
            self.start_date,
            self.end_date,
        ]:
            widget.setStyleSheet(control_style)

        # Export button style
        self.export_btn.setStyleSheet(
            """
            QPushButton {
                background-color: #00B4D8;
                color: white;
                border-radius: 5px;
                padding: 8px 15px;
            }
            QPushButton:hover { background-color: #0096B7; }
        """
        )
        
        self.export_format_combo.setStyleSheet(control_style)

        # Connect signals
        self.range_combo.currentIndexChanged.connect(self.update_date_range)
        self.view_combo.currentIndexChanged.connect(self.toggle_view)
        self.export_btn.clicked.connect(self.handle_export)
        self.status_combo.currentIndexChanged.connect(self.update_display)
        self.start_date.dateChanged.connect(self.update_display)
        self.end_date.dateChanged.connect(self.update_display)

    def toggle_view(self):
        """Switch between different views"""
        # UPDATED: Handle all view types
        current_view = self.view_combo.currentText()
        if current_view == "Text View":
            current_index = 1  # List widget
        else:
            current_index = 0  # Canvas for all graph types
        
        self.stacked_widget.setCurrentIndex(current_index)
        self.adjustSize()
        self.update_display()

    def update_date_range(self):
        if self.range_combo.currentText() == "Custom":
            self.start_date.setEnabled(True)
            self.end_date.setEnabled(True)
        else:
            today = QDate.currentDate()
            if self.range_combo.currentText() == "This Week":
                # Calculate the start of the week (Monday)
                # In Qt, Monday is 1, Tuesday is 2, etc.
                # We need to go back (dayOfWeek() - 1) days to get to Monday
                days_to_monday = today.dayOfWeek() - 1
                start = today.addDays(-days_to_monday)

                # Calculate the end of the week (Sunday)
                # Go forward 6 days from Monday to get to Sunday
                end = start.addDays(6)
            elif self.range_combo.currentText() == "This Month":
                start = QDate(today.year(), today.month(), 1)
                end = today
            else:  # This Year
                start = QDate(today.year(), 1, 1)
                end = today

            self.start_date.setDate(start)
            self.end_date.setDate(end)
            self.start_date.setEnabled(False)
            self.end_date.setEnabled(False)

        self.update_display()

    def update_display(self):

        start = self.start_date.date().toPyDate()
        end = self.end_date.date().toPyDate()

        status_filter = "all"
        match self.status_combo.currentText():
            case "Done":
                status_filter = "done"
            case "Failed":
                status_filter = "failed"

        # UPDATED: Handle all view types
        current_view = self.view_combo.currentText()
        if current_view == "Text View":
            self.update_text_history(start, end, status_filter)
        elif current_view == "Bar Graph":
            self.update_bar_graph(start, end, status_filter)
        elif current_view == "Line Graph":
            self.update_line_graph(start, end, status_filter)
        elif current_view == "Pie Chart":
            self.update_pie_chart(start, end, status_filter)

    def update_bar_graph(self, start, end, status_filter):
        """Original bar graph functionality"""
        # UPDATED: Add username filter
        done, failed, _ = HistoryManager.load_history(
            self.username, start, end, status_filter
        )

        dates = []
        current = start
        while current <= end:
            dates.append(current)
            current += timedelta(days=1)

        done_counts = [done.get(d.strftime("%Y-%m-%d"), 0) for d in dates]
        failed_counts = [failed.get(d.strftime("%Y-%m-%d"), 0) for d in dates]

        self.figure.clear()
        ax = self.figure.add_subplot(111)
        
        total_done = sum(done_counts)
        total_failed = sum(failed_counts)
        
        if total_done == 0 and total_failed == 0:
            ax.text(0.5, 0.5, 'No data available\nfor selected period', 
                   ha='center', va='center', transform=ax.transAxes, 
                   fontsize=16, color='gray')
            ax.axis('off')
        else:
            bar_width = 0.35  # Width of the bars

            if status_filter == "all":
                # Convert dates to numbers for proper bar positioning
                x = [mdates.date2num(d) for d in dates]

                # Plot bars side by side
                done_bars = ax.bar(
                    [xi - bar_width / 2 for xi in x],
                    done_counts,
                    bar_width,
                    label="Done",
                    color="#4CAF50",
                )
                failed_bars = ax.bar(
                    [xi + bar_width / 2 for xi in x],
                    failed_counts,
                    bar_width,
                    label="Failed",
                    color="#FF4444",
                )

                # Add value labels on top of bars
                for bars in [done_bars, failed_bars]:
                    for bar in bars:
                        height = bar.get_height()
                        if height > 0:  # Only show label if there's a value
                            ax.text(
                                bar.get_x() + bar.get_width() / 2,
                                height,
                                f"{int(height)}",
                                ha="center",
                                va="bottom",
                            )
            elif status_filter == "done":
                bars = ax.bar(dates, done_counts, color="#4CAF50")
                for bar in bars:
                    height = bar.get_height()
                    if height > 0:
                        ax.text(
                            bar.get_x() + bar.get_width() / 2,
                            height,
                            f"{int(height)}",
                            ha="center",
                            va="bottom",
                        )
            else:
                bars = ax.bar(dates, failed_counts, color="#FF4444")
                for bar in bars:
                    height = bar.get_height()
                    if height > 0:
                        ax.text(
                            bar.get_x() + bar.get_width() / 2,
                            height,
                            f"{int(height)}",
                            ha="center",
                            va="bottom",
                        )

        ax.set_title("Task Completion History - Bar Chart")
        
        ax.xaxis.set_major_locator(mdates.AutoDateLocator())
        if len(dates) <= 7:  # For short periods, show full dates
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
        else:  # For longer periods, show abbreviated
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%m-%d"))
        ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
        
        ax.set_ylim(bottom=0)
        self.figure.autofmt_xdate()
        self.canvas.draw()

    def update_line_graph(self, start, end, status_filter):
        """NEW: Line graph view"""
        done, failed, _ = HistoryManager.load_history(
            self.username, start, end, status_filter
        )

        dates = []
        current = start
        while current <= end:
            dates.append(current)
            current += timedelta(days=1)

        done_counts = [done.get(d.strftime("%Y-%m-%d"), 0) for d in dates]
        failed_counts = [failed.get(d.strftime("%Y-%m-%d"), 0) for d in dates]

        self.figure.clear()
        ax = self.figure.add_subplot(111)

        # Check if we have any data
        total_done = sum(done_counts)
        total_failed = sum(failed_counts)
        
        if total_done == 0 and total_failed == 0:
            ax.text(0.5, 0.5, 'No data available\nfor selected period', 
                   ha='center', va='center', transform=ax.transAxes, 
                   fontsize=16, color='gray')
            ax.axis('off')
        else:
            if status_filter == "all":
                # Plot both lines
                line1 = ax.plot(dates, done_counts, marker='o', linewidth=2.5, markersize=6, 
                       color="#4CAF50", label="Done", alpha=0.8)
                line2 = ax.plot(dates, failed_counts, marker='s', linewidth=2.5, markersize=6, 
                       color="#FF4444", label="Failed", alpha=0.8)
                
                # Add area fill for better visualization
                ax.fill_between(dates, done_counts, alpha=0.2, color="#4CAF50")
                ax.fill_between(dates, failed_counts, alpha=0.2, color="#FF4444")
                
                ax.legend(loc='upper right')
            elif status_filter == "done":
                ax.plot(dates, done_counts, marker='o', linewidth=2.5, markersize=6, 
                       color="#4CAF50", alpha=0.8)
                ax.fill_between(dates, done_counts, alpha=0.2, color="#4CAF50")
            else:
                ax.plot(dates, failed_counts, marker='s', linewidth=2.5, markersize=6, 
                       color="#FF4444", alpha=0.8)
                ax.fill_between(dates, failed_counts, alpha=0.2, color="#FF4444")

            # Improve styling
            ax.set_xlabel("Date", fontsize=11)
            ax.set_ylabel("Number of Tasks", fontsize=11)
            
            # Enhanced grid
            ax.grid(True, alpha=0.3, linestyle='--')
            ax.set_facecolor('#fafafa') 
            # Better date formatting
            ax.xaxis.set_major_locator(mdates.AutoDateLocator())
            if len(dates) <= 7:  # For short periods, show full dates
                ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
            else:  # For longer periods, show abbreviated
                ax.xaxis.set_major_formatter(mdates.DateFormatter("%m-%d"))
            
            ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
            
            # Set y-axis to start from 0 but add some padding at top
            max_val = max(max(done_counts) if done_counts else [0], 
                         max(failed_counts) if failed_counts else [0])
            if max_val > 0:
                ax.set_ylim(0, max_val * 1.1)  # Add 10% padding at top
            else:
                ax.set_ylim(0, 1)
        
        ax.set_title("Task Completion History - Line Chart")
        self.figure.autofmt_xdate()
        self.canvas.draw()

    def update_pie_chart(self, start, end, status_filter):
        """NEW: Pie chart view"""
        done, failed, _ = HistoryManager.load_history(
            self.username, start, end, status_filter
        )
        
        
        dates = []
        current = start
        while current <= end:
            dates.append(current)
            current += timedelta(days=1)

        done_counts = [done.get(d.strftime("%Y-%m-%d"), 0) for d in dates]
        failed_counts = [failed.get(d.strftime("%Y-%m-%d"), 0) for d in dates]
        
        # Calculate totals
        total_done = sum(done_counts)
        total_failed = sum(failed_counts)

        self.figure.clear()
        ax = self.figure.add_subplot(111)

        if status_filter == "all":
            if total_done == 0 and total_failed == 0:
                # No data case
                ax.text(0.5, 0.5, 'No data available\nfor selected period', 
                       ha='center', va='center', transform=ax.transAxes, 
                       fontsize=16, color='gray')
                ax.axis('off')
            else:
                sizes = []
                labels = []
                colors = []
                
                if total_done > 0:
                    sizes.append(total_done)
                    labels.append(f'Done ({total_done})')
                    colors.append('#4CAF50')
                
                if total_failed > 0:
                    sizes.append(total_failed)
                    labels.append(f'Failed ({total_failed})')
                    colors.append('#FF4444')
                
                wedges, texts, autotexts = ax.pie(sizes, labels=labels, colors=colors, 
                                                 autopct='%1.1f%%', startangle=90)
                
                # Make percentage text bold and white
                for autotext in autotexts:
                    autotext.set_color('white')
                    autotext.set_fontweight('bold')
                    autotext.set_fontsize(10)
        else:
            # Single status filter
            if status_filter == "done":
                total = total_done
                color = '#4CAF50'
                status_name = "Done"
            else:
                total = total_failed
                color = '#FF4444'
                status_name = "Failed"
            
            if total == 0:
                ax.text(0.5, 0.5, f'No {status_name.lower()} tasks\nfor selected period', 
                       ha='center', va='center', transform=ax.transAxes, 
                       fontsize=16, color='gray')
                ax.set_xlim(0, 1)
                ax.set_ylim(0, 1)
            else:
                # Show daily breakdown for single status
                daily_counts = []
                daily_labels = []
                
                current = start
                while current <= end:
                    date_str = current.strftime("%Y-%m-%d")
                    if status_filter == "done":
                        count = done.get(date_str, 0)
                    else:
                        count = failed.get(date_str, 0)
                    
                    if count > 0:
                        daily_counts.append(count)
                        daily_labels.append(current.strftime("%m-%d"))
                    
                    current += timedelta(days=1)
                
                if daily_counts:
                    wedges, texts, autotexts = ax.pie(daily_counts, labels=daily_labels, 
                                                     autopct='%1.1f%%', startangle=90)
                    for autotext in autotexts:
                        autotext.set_color('white')
                        autotext.set_fontweight('bold')
                        autotext.set_fontsize(8)
                else:
                    ax.text(0.5, 0.5, f'No {status_name.lower()} tasks\nfor selected period', 
                           ha='center', va='center', transform=ax.transAxes, 
                           fontsize=16, color='gray')
                    ax.set_xlim(0, 1)
                    ax.set_ylim(0, 1)

        ax.set_title("Task Completion History - Pie Chart")
        self.canvas.draw()


    def update_text_history(self, start, end, status_filter):
        # Load history with the correct date range
        _, _, entries = HistoryManager.load_history(
            self.username, start, end, status_filter
        )
        self.history_list.clear()

        # Style the QListWidget itself
        self.history_list.setStyleSheet(
            """
            QListWidget {
                background-color: #f5f5f5;  /* Light gray background */
                border-radius: 8px;
                padding: 8px;
            }
            QListWidget::item {
                background-color: white;     /* White background for items */
                border-radius: 4px;
                margin: 2px;
                padding: 8px;
            }
            QListWidget::item:hover {
                background-color: #e8f4f8;   /* Light blue on hover */
            }
        """
        )

        # Add a separator item at the beginning
        self.history_list.addItem("")

        # Filter entries by date range
        filtered_entries = []
        for entry in entries:
            status_date_str = None
            if "on" in entry["status"]:
                status_date_str = entry["status"].split("on ")[-1].strip()
            else:
                status_date_str = entry["status"].strip()

            try:
                # Parse the date
                status_date = datetime.strptime(status_date_str, "%Y-%m-%d").date()

                # Check if the date is within the selected range
                if start <= status_date <= end:
                    filtered_entries.append(entry)
            except ValueError:
                # Skip entries with invalid date formats
                continue

        # Display filtered entries
        for entry in filtered_entries:
            status = "🔴 Failed" if "failed" in entry["status"].lower() else "🟢 Done"
            text = (
                f"Task: {entry['task']}\n"
                f"Description: {entry['description']}\n"
                f"Start: {entry['start_time']} | Deadline: {entry['deadline']}\n"
                f"Priority: {entry['priority']} | Status: {status}"
            )
            item = QListWidgetItem(text)
            self.history_list.addItem(item)
            # Add a separator item after each task
            self.history_list.addItem("")

    # Update 19 Mei 2025
    # makeing option for export
    def handle_export(self):
        format = self.export_format_combo.currentText()
        if format == "PDF":
            self.export_pdf()
        else:
            self.export_excel()

    def export_pdf(self):
        options = QFileDialog.Options()
        filename, _ = QFileDialog.getSaveFileName(
            self, "Save PDF", "", "PDF Files (*.pdf)", options=options
        )

        if filename:
            if not filename.endswith(".pdf"):
                filename += ".pdf"

            try:
                with PdfPages(filename) as pdf:
                    pdf.savefig(self.figure)
                QMessageBox.information(self, "Success", "PDF exported successfully!")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Export failed: {str(e)}")

    # update 19 Mei 2025
    # Export to Excel
    def export_excel(self):
        try:
            # Get date range and filter
            start = self.start_date.date().toPyDate()
            end = self.end_date.date().toPyDate()
            status_filter = "all"
            match self.status_combo.currentText():
                case "Done":
                    status_filter = "done"
                case "Failed":
                    status_filter = "failed"

            # Retrieve data
            _, _, entries = HistoryManager.load_history(
                self.username, start, end, status_filter
            )
            if not entries:
                QMessageBox.warning(self, "No Data", "No data to export.")
                return

            # Create DataFrame
            df = pd.DataFrame(entries)
            df["Status Type"] = df["status"].apply(
                lambda x: "Failed" if "failed" in x.lower() else "Done"
            )

            # File dialog
            options = QFileDialog.Options()
            filename, _ = QFileDialog.getSaveFileName(
                self, "Save Excel File", "", "Excel Files (*.xlsx)", options=options
            )
            if not filename:
                return

            if not filename.endswith(".xlsx"):
                filename += ".xlsx"

            # Save to Excel
            df.to_excel(filename, index=False, engine="openpyxl")
            wb = load_workbook(filename)
            ws = wb.active

            # Set widths for specific columns (example mapping)
            column_widths = {
                "task": 30,        # Wider for task names
                "description": 40, # Extra wide for descriptions
                "start_time": 17,  # Fixed width for start time
                "deadline": 17,    # Fixed width for deadline
                "status": 30,      # Fixed width for status
                "priority": 10     # Narrow for priority
            }

            # Apply widths to columns
            for idx, column_name in enumerate(df.columns, 1):
                column_letter = get_column_letter(idx)
                if column_name.lower() in column_widths:
                    ws.column_dimensions[column_letter].width = column_widths[column_name.lower()]
                else:
                    ws.column_dimensions[column_letter].width = 12  # Default width

            # Define colors
            header_fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"  # Blue
            )
            done_fill = PatternFill(
                start_color="A8D08D", end_color="A8D08D", fill_type="solid"  # Green
            )
            failed_fill = PatternFill(
                start_color="FF0000", end_color="FF0000", fill_type="solid"  # Red
            )

            # Color headers
            for cell in ws[1]:  # First row is headers
                cell.fill = header_fill

            # Find "Status Type" column index
            status_col_idx = df.columns.get_loc("Status Type") + 1  # +1 for Excel's 1-based index

            # Color rows based on status
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Start from row 2
                status_cell = ws.cell(row=row_idx, column=status_col_idx)
                fill = failed_fill if status_cell.value == "Failed" else done_fill

                for cell in row:
                    cell.fill = fill

            wb.save(filename)
            QMessageBox.information(self, "Success", "Excel file exported successfully!")

        except ImportError:
            QMessageBox.critical(
                self,
                "Error",
                "Required libraries (pandas/openpyxl) not installed.",
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Export failed: {str(e)}")